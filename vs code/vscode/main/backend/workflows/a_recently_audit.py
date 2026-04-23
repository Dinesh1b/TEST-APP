import sys
import os
import openpyxl
from typing import List, Dict, Optional
from dataclasses import dataclass
from enum import Enum
from playwright.sync_api import sync_playwright, Page, TimeoutError as PlaywrightTimeoutError
from pytest import Config



# ======================================================
# COLUMN INDICES (based on actual table structure)
# ======================================================
# 0:  S.No
# 1:  Item Code
# 2:  Item Name
# 3:  UOM
# 4:  Auditor Name
# 5:  Audited Qty
# 6:  Damage Qty
# 7:  Audited Image   ← image tag presence check, not text
# 8:  Stock Qty
# 9:  Audited Location

COL_S_NO            = 0
COL_ITEM_CODE       = 1
COL_ITEM_NAME       = 2
COL_UOM             = 3
COL_AUDITOR         = 4
COL_AUDITED_QTY     = 5
COL_DAMAGE_QTY      = 6
COL_AUDITED_IMAGE   = 7   # presence/absence of <img> tag
COL_STOCK_QTY       = 8
COL_AUDITED_LOCATION= 9


# ======================================================
# VALIDATION RESULT CLASSES
# ======================================================
class ValidationStatus(Enum):
    PASSED  = "PASSED"
    FAILED  = "FAILED"
    WARNING = "WARNING"
    SKIPPED = "SKIPPED"


@dataclass
class FieldValidation:
    field_name:     str
    expected_value: str
    actual_value:   str
    status:         ValidationStatus
    error_message:  Optional[str] = None

    @property
    def passed(self) -> bool:
        return self.status == ValidationStatus.PASSED

    def __str__(self) -> str:
        # Always show Expected → Actual so every field is fully traceable
        return (f"  {self.status.value} {self.field_name}: "
                f"Expected='{self.expected_value}'  →  Got='{self.actual_value}'")


@dataclass
class RowValidation:
    row_number:        int
    item_code:         str
    field_validations: List[FieldValidation]
    overall_status:    ValidationStatus
    error_message:     Optional[str] = None

    @property
    def passed(self) -> bool:
        return all(fv.passed for fv in self.field_validations)

    @property
    def failed_fields(self) -> List[FieldValidation]:
        return [fv for fv in self.field_validations if not fv.passed]

    def __str__(self) -> str:
        header = f"\nRow {self.row_number} ({self.item_code}): {self.overall_status.value}"
        if self.error_message:
            header += f" - {self.error_message}"
        field_results = "\n".join(str(fv) for fv in self.field_validations)
        return f"{header}\n{field_results}"


@dataclass
class ValidationReport:
    total_rows_expected: int
    total_rows_actual:   int
    row_validations:     List[RowValidation]
    overall_status:      ValidationStatus
    summary:             Dict[str, int]

    @property
    def passed_rows(self) -> List[RowValidation]:
        return [rv for rv in self.row_validations if rv.passed]

    @property
    def failed_rows(self) -> List[RowValidation]:
        return [rv for rv in self.row_validations if not rv.passed]

    def print_report(self):
        """
        Clean, focused report output - aligned columns, failures only.
        """
        print("\n" + "=" * 100)
        print("VALIDATION REPORT")
        print("=" * 100)

        # ── EXECUTIVE SUMMARY ────────────────────────────────────
        status_emoji = "✅" if self.overall_status == ValidationStatus.PASSED else "❌"
        print(f"\n{status_emoji} OVERALL: {self.overall_status.value}")
        print(f"  Expected Rows:  {self.total_rows_expected}")
        print(f"  Actual Rows:    {self.total_rows_actual}")
        print(f"  Passed:         {len(self.passed_rows)}")
        print(f"  Failed:         {len(self.failed_rows)}")

        # ── ROW COUNT MISMATCH (if applicable) ────────────────────────────────────
        if self.total_rows_expected != self.total_rows_actual:
            mismatch_indicator = "⚠️ " if self.total_rows_actual > self.total_rows_expected else "❌"
            print(f"\n{mismatch_indicator} ROW COUNT MISMATCH")
            print(f"  Expected {self.total_rows_expected} rows, found {self.total_rows_actual}")

        # ── FAILURES ONLY (skip if all passed) ───────────────────────────────────
        if self.failed_rows:
            print("\n" + "-" * 100)
            print(f"FAILURES ({len(self.failed_rows)} row(s))")
            print("-" * 100)

            for rv in self.failed_rows:
                # Row header with item code for quick identification
                header_line = f"\nRow {rv.row_number} | {rv.item_code}"
                if rv.error_message:
                    header_line += f" | {rv.error_message}"
                print(header_line)

                # Show only failed fields (aligned columns with status icons)
                if rv.failed_fields:
                    # Calculate column widths
                    max_field_len = max(len(fv.field_name) for fv in rv.failed_fields)
                    max_expected_len = max(len(fv.expected_value) for fv in rv.failed_fields)

                    for fv in rv.failed_fields:
                        status_icon = "✅" if fv.passed else "❌"
                        field_col = fv.field_name.ljust(max_field_len)
                        expected_col = fv.expected_value.ljust(max_expected_len)
                        print(f"  {status_icon} {field_col}  →  Expected: {expected_col}  |  Actual: {fv.actual_value}")
        else:
            print("\n✅ All rows passed validation")

        # ── STATISTICS ──────────────────────────────────────
        if self.row_validations:
            total_fields  = sum(len(rv.field_validations) for rv in self.row_validations)
            passed_fields = sum(sum(1 for fv in rv.field_validations if fv.passed)
                                for rv in self.row_validations)
            failed_fields = total_fields - passed_fields

            accuracy = (passed_fields / total_fields) * 100 if total_fields > 0 else 0

            print("\n" + "-" * 100)
            print("STATISTICS")
            print("-" * 100)
            print(f"  Total Fields:  {total_fields}")
            print(f"  Passed:        {passed_fields}")
            print(f"  Failed:        {failed_fields}")
            print(f"  Accuracy:      {accuracy:.1f}%")

        print("\n" + "=" * 100 + "\n")

    # ── Timing / polling ────────────────────────────────
    TABLE_LOAD_TIMEOUT    = 60000
    ROW_LOAD_TIMEOUT      = 20000
    NETWORK_IDLE_TIMEOUT  = 15000
    POLLING_INTERVAL      = 1000
    MAX_POLL_ATTEMPTS     = 60


# ======================================================
# UTILITY FUNCTIONS
# ======================================================
def normalize(text: str) -> str:
    return text.replace(",", "").strip() if text else ""


def compare_values(expected: str, actual: str, field_name: str) -> FieldValidation:
    if normalize(expected) == normalize(actual):
        return FieldValidation(field_name, expected, actual, ValidationStatus.PASSED)
    return FieldValidation(field_name, expected, actual, ValidationStatus.FAILED,
                        "Value mismatch")


def is_fully_blank_row(row: Dict[str, str]) -> bool:
    keys_to_check = [
        "item_code",
        "item_name",
        "uom",
        "auditor_name",
        "audited_qty",
        "damaged_qty",
        "stock_qty",
        "audited_location",
    ]
    return all(not row.get(k, "").strip() for k in keys_to_check)

def load_expected_rows_from_excel(
        excel_path: str,
        sheet_name: Optional[str] = None
    ) -> List[Dict[str, str]]:
    """
    Read Recently_Audited.xlsx and return a list of dicts matching
    the keys used by the validator:
        item_code, item_name, uom, auditor_name,
        audited_qty, damaged_qty, stock_qty, audited_location
    Row 1 is treated as the header; data starts at row 2.
    """
    if not os.path.exists(excel_path):
        print(f"❌ ERROR: Excel file not found at {excel_path}")
        return []

    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # ✅ Sheet selection
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            print(f"❌ ERROR: Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")
            return []
        ws = wb[sheet_name]
        print(f"📄 Loaded Excel sheet: {sheet_name}")
    else:
        ws = wb.active
        print(f"📄 Loaded Excel active sheet: {ws.title}")
    rows = list(ws.iter_rows(values_only=True))
    print(len(rows))
    # Map header → column index
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    col_map = {h: i for i, h in enumerate(header)}

    expected = []
    for row in rows[1:]:

        def get(col_name: str) -> str:
            idx = col_map.get(col_name)
            if idx is None:
                return ""
            val = row[idx]
            if isinstance(val, float) and val.is_integer():
                return str(int(val))
            return str(val).strip() if val is not None else ""

        row_data = {
            "s_no":             get("S.No"),
            "item_code":        get("Item Code"),
            "item_name":        get("Item Name"),
            "uom":              get("UOM"),
            "auditor_name":     get("Auditor Name"),
            "audited_qty":      get("Audited Qty"),
            "damaged_qty":      get("Damage Qty"),
            "stock_qty":        get("Stock Qty"),
            "audited_location": get("Audited Location"),
            "audited_image":    "",
        }

        # 🚫 SKIP fully blank Excel rows (ignore S.No)
        if all(not value.strip() for key, value in row_data.items() if key != "s_no"):
            
            continue

        expected.append(row_data)
    return expected
# ======================================================
# TABLE SELECTORS & WAITING
# ======================================================
def try_multiple_table_selectors(page: Page) -> Optional[any]:
    print("\n🎯 Trying multiple table selectors...")
    selectors = [
        ("xpath=//*[contains(text(),'Recently Audited')]/following::table[1]", "XPath - following table"),
        ("xpath=//div[contains(text(),'Recently Audited')]/following-sibling::*//table", "XPath - sibling table"),
        (".common-table table",        "CSS - common-table class"),
        ("p-table table",              "CSS - p-table element"),
        (".p-datatable-wrapper table", "CSS - datatable wrapper"),
        ("table >> nth=0", "First table"),
        ("table >> nth=1", "Second table"),
        ("table >> nth=2", "Third table"),
    ]
    for selector, description in selectors:
        try:
            print(f"  Trying: {description}")
            table = page.locator(selector)
            if table.count() > 0 and table.locator("tbody").first.count() > 0:
                print(f"    ✅ Found table using: {description}")
                return table.first
            print("    ❌ No match")
        except Exception as e:
            print(f"    ❌ Error: {e}")
    print("  ❌ No suitable table found")
    return None


def wait_for_table_with_polling(page: Page) -> Optional[any]:
    print(f"\n⏳ Polling for table (max 60 attempts)...")
    for attempt in range(60):
        try:
            if attempt % 5 == 0:
                print(f"  Attempt {attempt + 1}/60...")
            page.wait_for_load_state("domcontentloaded", timeout=2000)
            table = try_multiple_table_selectors(page)
            if table and table.locator("tbody tr").count() > 0:
                print(f"\n  ✅ Found table with {table.locator('tbody tr').count()} rows")
                return table
            page.wait_for_timeout(1000)
            if attempt % 5 == 0:
                page.keyboard.press("PageDown")
                page.wait_for_timeout(500)
        except Exception as e:
            if attempt % 20 == 0:
                print(f"    ⚠️  Polling error: {str(e)[:100]}")
    print("  ❌ Polling timeout")
    return None


def wait_for_table_to_load(page: Page) -> bool:
    print("\n" + "=" * 70)
    print("⏳ WAITING FOR TABLE TO LOAD")
    print("=" * 70)
    try:
        try:
            page.wait_for_load_state("networkidle", timeout=15000)
            print("  ✅ Network idle")
        except:
            print("  ⚠️  Network idle timeout (continuing)")
        page.wait_for_load_state("domcontentloaded", timeout=5000)
        page.wait_for_timeout(5000)
        page.keyboard.press("PageDown")
        page.wait_for_timeout(2000)
        return wait_for_table_with_polling(page) is not None
    except Exception as e:
        print(f"\n❌ ERROR: {e}")
        return False


# ======================================================
# ROW EXTRACTION (FIXED FOR TEXT WRAPPING)
# ======================================================
def extract_row_data(row, row_index: int) -> Dict[str, str]:
    """
    Extract row data with improved text handling.
    Uses text_content() instead of inner_text() to capture wrapped text
    and normalize whitespace from wrapped elements.
    """
    cells = row.locator("td")

    def safe_text(col_idx: int) -> str:
        """Extract text, handling text wrapping by collapsing whitespace."""
        try:
            # text_content() captures all text including wrapped text
            text = cells.nth(col_idx).text_content(timeout=5000).strip()
            # Remove extra whitespace and newlines from text wrapping
            text = " ".join(text.split())
            return text
        except:
            return "N/A"

    def has_image(col_idx: int) -> str:
        """Return 'Yes' if cell contains a real <img>, else 'No'."""
        try:
            img = cells.nth(col_idx).locator("img")
            if img.count() > 0:
                src = img.first.get_attribute("src", timeout=3000) or ""
                if src and "no_image" not in src:
                    return "Yes"
            return "No"
        except:
            return "No"

    return {
        "s_no":             safe_text(COL_S_NO),               # col 0
        "item_code":        safe_text(COL_ITEM_CODE),          # col 1
        "item_name":        safe_text(COL_ITEM_NAME),          # col 2
        "uom":              safe_text(COL_UOM),                # col 3
        "auditor_name":     safe_text(COL_AUDITOR),            # col 4
        "audited_qty":      safe_text(COL_AUDITED_QTY),        # col 5
        "damaged_qty":      safe_text(COL_DAMAGE_QTY),         # col 6
        "audited_image":    has_image(COL_AUDITED_IMAGE),      # col 7 — img presence
        "stock_qty":        safe_text(COL_STOCK_QTY),          # col 8
        "audited_location": safe_text(COL_AUDITED_LOCATION),   # col 9 — handles wrapped text
    }


# ======================================================
# VALIDATION ENGINE
# ======================================================
def validate_row_against_expected(
        actual: Dict[str, str],
        expected: Dict[str, str],
    ) -> RowValidation:
    """
    Validate one actual row against its item_code-matched expected row.
    S.No is completely ignored — order is irrelevant.
    """
    validations: List[FieldValidation] = []
 
    fields = [
        ("Item Name",        "item_name"),
        ("UOM",              "uom"),
        ("Auditor Name",     "auditor_name"),
        ("Audited Qty",      "audited_qty"),
        ("Damage Qty",       "damaged_qty"),
        ("Stock Qty",        "stock_qty"),
        ("Audited Location", "audited_location"),
    ]
    for label, key in fields:
        validations.append(
            compare_values(expected.get(key, ""), actual.get(key, ""), label)
        )
 
    # Audited Image — SKIPPED when Excel has no expected value
    img_actual   = actual.get("audited_image", "No")
    img_expected = expected.get("audited_image", "")
    if img_expected == "":
        validations.append(FieldValidation(
            field_name="Audited Image",
            expected_value="(not in Excel)",
            actual_value=img_actual,
            status=ValidationStatus.SKIPPED,
        ))
    else:
        validations.append(compare_values(img_expected, img_actual, "Audited Image"))
 
    non_skipped = [v for v in validations if v.status != ValidationStatus.SKIPPED]
    status = (ValidationStatus.PASSED if all(v.passed for v in non_skipped)
              else ValidationStatus.FAILED)
 
    return RowValidation(
        row_number=0,                        # position irrelevant
        item_code=actual.get("item_code", "?"),
        field_validations=validations,
        overall_status=status,
    )
 
 
def validate_excel_columns_only(
        rows_data: List[Dict[str, str]],
        expected_data: List[Dict[str, str]],
    ) -> ValidationReport:
    """
    Item-code-keyed validation. Row order in UI vs Excel is completely irrelevant.
 
    actual item_code not in Excel  -> WARNING  (extra / unexpected row)
    Excel item_code not in UI      -> FAILED   (missing row)
    """
    row_validations: List[RowValidation] = []
 
    # Build lookup: item_code -> expected row
    expected_by_code: Dict[str, Dict[str, str]] = {}
    for exp in expected_data:
        code = exp.get("item_code", "").strip()
        if code and code not in expected_by_code:
            expected_by_code[code] = exp
 
    matched_codes: set = set()
 
    # Validate each scraped UI row
    for actual in rows_data:
        code = actual.get("item_code", "").strip()
        exp  = expected_by_code.get(code)
 
        if exp is None:
            row_validations.append(RowValidation(
                row_number=0,
                item_code=code or "?",
                field_validations=[],
                overall_status=ValidationStatus.WARNING,
                error_message=f"'{code}' in UI but NOT in Excel",
            ))
        else:
            matched_codes.add(code)
            row_validations.append(validate_row_against_expected(actual, exp))
 
    # Flag Excel rows missing from UI
    for exp in expected_data:
        code = exp.get("item_code", "").strip()
        if code and code not in matched_codes:
            row_validations.append(RowValidation(
                row_number=0,
                item_code=code,
                field_validations=[],
                overall_status=ValidationStatus.FAILED,
                error_message=f"'{code}' in Excel but NOT found in UI",
            ))
 
    counts_match = len(rows_data) == len(expected_data)
    all_passed   = all(
        rv.overall_status in (ValidationStatus.PASSED, ValidationStatus.WARNING)
        for rv in row_validations
    )
    overall = ValidationStatus.PASSED if all_passed and counts_match else ValidationStatus.FAILED
 
    return ValidationReport(
        total_rows_expected=len(expected_data),
        total_rows_actual=len(rows_data),
        row_validations=row_validations,
        overall_status=overall,
        summary={
            "passed":  sum(1 for rv in row_validations if rv.overall_status == ValidationStatus.PASSED),
            "failed":  sum(1 for rv in row_validations if rv.overall_status == ValidationStatus.FAILED),
            "warning": sum(1 for rv in row_validations if rv.overall_status == ValidationStatus.WARNING),
        }
    )
 

# ======================================================
# PAGINATION HELPERS
# ======================================================
def get_total_entries(page: Page) -> int:
    """Read total row count from 'Showing X to Y of Z entries' text."""
    try:
        text = page.locator("#pageshow").inner_text(timeout=5000).strip()
        # e.g. "Showing 1 to 5 of 12 entries"
        total = int(text.split("of")[1].split("entries")[0].strip())
        print(f"  📄 Paginator says: '{text}' → total entries: {total}")
        return total
    except Exception as e:
        print(f"  ⚠️  Could not read total entries: {e}")
        return -1


def extract_current_page_rows(page: Page, global_offset: int) -> List[Dict[str, str]]:
    """Extract all visible rows on the current page with wrapped text support."""
    table = try_multiple_table_selectors(page)
    if not table:
        return []
    rows = table.locator("tbody tr")
    count = rows.count()
    results = []
    
    col_widths = {
        "s_no": 4,
        "item_code": 10,
        "item_name": 12,
        "uom": 4,
        "auditor_name": 10,
        "audited_qty": 8,
        "damaged_qty": 8,
        "audited_image": 5,
        "stock_qty": 6,
        "audited_location": 30,  # Increased for wrapped location text
    }
    
    for i in range(count):
        try:
            data = extract_row_data(rows.nth(i), global_offset + i)
            # 🚫 Skip fully blank UI rows
            if is_fully_blank_row(data):
                print(f"  ⏭️  Skipping blank UI row at position {global_offset + i + 1}")
                continue
            results.append(data)
            
            # Build aligned row output (with location text truncated for display)
            s_no_col = data['s_no'].ljust(col_widths['s_no'])
            code_col = data['item_code'].ljust(col_widths['item_code'])
            name_col = data['item_name'][:12].ljust(col_widths['item_name'])
            uom_col = data['uom'].ljust(col_widths['uom'])
            auditor_col = data['auditor_name'][:10].ljust(col_widths['auditor_name'])
            audited_col = data['audited_qty'].ljust(col_widths['audited_qty'])
            damaged_col = data['damaged_qty'].ljust(col_widths['damaged_qty'])
            image_col = data['audited_image'].ljust(col_widths['audited_image'])
            stock_col = data['stock_qty'].ljust(col_widths['stock_qty'])
            location_col = data['audited_location'][:30]  # Show up to 30 chars
            
            print(f"  {s_no_col} | {code_col} | {name_col} | {uom_col} | {auditor_col} | "
                f"{audited_col} | {damaged_col} | {image_col} | {stock_col} | {location_col}")
        except Exception as e:
            print(f"  ❌ Error extracting row {global_offset + i + 1}: {e}")
    return results


def go_to_next_page(page: Page) -> bool:
    """Click the Next page button. Returns True if clicked, False if disabled."""
    try:
        next_btn = page.locator("button.p-paginator-next")
        if next_btn.count() == 0:
            return False
        classes = next_btn.get_attribute("class") or ""
        if "p-disabled" in classes:
            return False
        next_btn.click()
        page.wait_for_load_state("networkidle", timeout=10000)
        page.wait_for_timeout(1500)
        return True
    except Exception as e:
        print(f"  ⚠️  Next page error: {e}")
        return False


# ======================================================
# COMPREHENSIVE VALIDATION
# ======================================================
def validate_table_comprehensive(
        page: Page,
        expected_rows: List[Dict[str, str]]
    ) -> ValidationReport:
    print("\n" + "=" * 80)
    print("🔍 COMPREHENSIVE TABLE VALIDATION  (all pages)")
    print("=" * 80)

    if not wait_for_table_to_load(page):
        return ValidationReport(len(expected_rows), 0, [], ValidationStatus.FAILED,
                                {"error": "Table not found"})

    total_entries = get_total_entries(page)
    print(f"\n📊 Total entries in UI: {total_entries} | Expected from Excel: {len(expected_rows)}")

    # ── Scrape every page ────────────────────────────────────────────────────
    all_rows_data: List[Dict[str, str]] = []
    current_page = 1

    while True:
        print(f"\n📄 Scraping page {current_page}...")
        
        # Print header on first page
        if current_page == 1:
            print("  S.NO | ITEM CODE  | ITEM NAME    | UOM | AUDITOR    | AUDITED_QTY | DAMAGED_QTY | IMAGE | STOCK |           LOCATION")
            print("  -----|------------|--------------|-----|------------|-------------|-------------|-------|-------|------------------------------")
        
        page_rows = extract_current_page_rows(page, global_offset=len(all_rows_data))
        all_rows_data.extend(page_rows)

        # Try to advance to next page
        if not go_to_next_page(page):
            print(f"\n  ✅ No more pages — scraped {len(all_rows_data)} rows total")
            break
        current_page += 1

    # ── Validate against Excel ───────────────────────────────────────────────
    return validate_excel_columns_only(all_rows_data, expected_rows)

    # --------------------------------------------------------------
    # SELECT BRANCH using XPATH (PrimeNG p-dropdown)
    # --------------------------------------------------------------
def select_branch(page, branch_text: str):

        # 1️⃣ Branch dropdown container (XPath)
        dropdown = page.locator(
            'xpath=/html/body/app-root/app-layout/div/app-header/div[1]/div[2]/div[2]//p-dropdown'
        )
        dropdown.wait_for(state="visible", timeout=30000)

        # 2️⃣ Click dropdown label to open
        dropdown.locator(".p-dropdown-label").click()

        # 3️⃣ Wait for dropdown panel (overlay)
        panel = page.locator(".p-dropdown-panel").last
        panel.wait_for(state="visible", timeout=30000)

        # 4️⃣ Filter search (if exists)
        search = panel.locator("input.p-dropdown-filter")
        if search.count() > 0:
            search.fill("")
            search.fill(branch_text)
            page.wait_for_timeout(300)

        # 5️⃣ Select branch option
        option = panel.locator(
            "li.p-dropdown-item",
            has_text=branch_text
        )

        if option.count() == 0:
            raise Exception(f"❌ Branch not found: {branch_text}")

        option.first.click()

        # 6️⃣ Ensure panel closed
        panel.wait_for(state="hidden", timeout=30000)

        # 7️⃣ Final verification
        dropdown.locator(
            ".p-dropdown-label",
            has_text=branch_text
        ).wait_for(state="visible", timeout=30000)

# ======================================================
# NAVIGATION
# ======================================================
def navigate_to_ongoing_recentlyaudit(page, config):
    """Navigate to the ongoing audits Page and open specified audit"""

    print("📂 Navigating to Ongoing Audits...")
    # Wait until element visible
    page.wait_for_selector("a[href='/home/audit']", state="visible")

    # Click
    page.click("a[href='/home/audit']")
    page.wait_for_load_state("networkidle")

    select_branch(page, config.Branch)

    print(f"✅ Branch selected: {config.Branch}")
    # Find and click the specific audit
    row = page.locator(
        f"table tbody tr:has(td span:has-text('{config.ap_audit_name}'))"
    ).first

    row.locator("td span").first.click()
    page.wait_for_load_state("networkidle")
    print(f"✅ Opened audit: {config.ap_audit_name}")


# ======================================================
# MAIN VALIDATION FUNCTION
# ======================================================
def A_Recently_Audit(page, Config):
    """
    Main entry point for Recently Audited table validation.
    """
    print("\n" + "=" * 80)
    print("🚀 RECENTLY AUDITED TABLE VALIDATION")
    print("=" * 80 + "\n")

    # Navigate to the audit
    navigate_to_ongoing_recentlyaudit(page, Config)

    # Load expected data from Excel
    expected_rows = load_expected_rows_from_excel(
    Config.EXCEL_PATH,
    sheet_name=Config.EXCEL_SHEET_Recently_Audit
)   
    print(len(expected_rows))
    print(f"📂 Loaded {len(expected_rows)} expected rows from Excel:\n")
    
    if expected_rows:
        # Get all unique keys from all rows
        all_keys = set()
        for row in expected_rows:
            all_keys.update(row.keys())
        
        # Define column order
        column_order = [
            "s_no", "item_code", "item_name", "uom", "auditor_name",
            "audited_qty", "damaged_qty", "stock_qty", "audited_location", "audited_image"
        ]
        keys_to_display = [k for k in column_order if k in all_keys]
        
        # Calculate column widths
        col_widths = {key: len(key) for key in keys_to_display}
        for row in expected_rows:
            for key in keys_to_display:
                col_widths[key] = max(col_widths[key], len(str(row.get(key, ""))))
        
        # Print header
        header_parts = []
        for key in keys_to_display:
            header_parts.append(key.upper().ljust(col_widths[key]))
        print("  " + " | ".join(header_parts))
        print("  " + "-" * (sum(col_widths.values()) + (len(keys_to_display) - 1) * 3))
        
        # Print rows
        for i, row in enumerate(expected_rows, 1):
            row_parts = []
            for key in keys_to_display:
                value = str(row.get(key, "")).strip()
                row_parts.append(value.ljust(col_widths[key]))
            print("  " + " | ".join(row_parts))

        try:
            print("\n⏳ Waiting 10 seconds for page to stabilize...")
            page.wait_for_timeout(10000)

            report = validate_table_comprehensive(page, expected_rows)
            report.print_report()

        except Exception as e:
            print(f"\n❌ CRITICAL ERROR: {e}")
            import traceback
            traceback.print_exc()

    print("\n✅ VALIDATION COMPLETED\n")