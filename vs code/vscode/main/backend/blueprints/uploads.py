"""
uploads.py — File upload and Excel inspection routes.
"""

import os
import base64
from io import BytesIO

from flask import Blueprint, request, jsonify, send_file
from openpyxl import load_workbook

uploads_bp = Blueprint("uploads", __name__)

_BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
UPLOAD_DIR = os.path.join(_BASE, "backend", "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

_ALLOWED_UPLOAD_EXTS = {".xlsx", ".xls"}


@uploads_bp.route("/upload", methods=["POST"])
def upload():
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify({"error": "No file provided", "path": ""}), 400

    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in _ALLOWED_UPLOAD_EXTS:
        return jsonify({"error": f"File type '{ext}' not allowed. Use .xlsx or .xls", "path": ""}), 400

    safe_name = os.path.basename(f.filename)
    dest = os.path.join(UPLOAD_DIR, safe_name)
    f.save(dest)
    return jsonify({"path": dest})


@uploads_bp.route("/inspect_audit_excel", methods=["POST"])
def inspect_audit_excel():
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify({"ok": False, "errors": ["No file provided."], "sheet_names": [], "headers_by_sheet": {}}), 400

    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in _ALLOWED_UPLOAD_EXTS:
        return jsonify({"ok": False, "errors": [f"File type '{ext}' not allowed. Use .xlsx or .xls"], "sheet_names": [], "headers_by_sheet": {}}), 400

    if ext == ".xls":
        return jsonify({
            "ok": False,
            "errors": [".xls inspection is not available on this machine. Please use .xlsx or enter the mapping manually."],
            "sheet_names": [],
            "headers_by_sheet": {},
        }), 400

    try:
        file_bytes = f.read()
        wb = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
        headers_by_sheet = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if row:
                headers_by_sheet[sheet_name] = [str(val).strip() for val in row if val is not None and str(val).strip()]
            else:
                headers_by_sheet[sheet_name] = []
        wb.close()
        default_sheet = "auditor_1" if "auditor_1" in headers_by_sheet else (wb.sheetnames[0] if wb.sheetnames else "")
        warnings = []
        if default_sheet and not headers_by_sheet.get(default_sheet):
            warnings.append(f"Sheet '{default_sheet}' has no detected header row.")
        return jsonify({
            "ok": True,
            "sheet_names": wb.sheetnames,
            "default_sheet": default_sheet,
            "headers_by_sheet": headers_by_sheet,
            "warnings": warnings,
            "errors": [],
        })
    except Exception as exc:
        return jsonify({
            "ok": False,
            "errors": [f"Unable to inspect the uploaded Excel: {exc}"],
            "sheet_names": [],
            "headers_by_sheet": {},
        }), 400


@uploads_bp.route("/sample_excel")
def sample_excel():
    import io
    from backend.sample_excel_b64 import _SAMPLE_EXCEL_B64
    data = base64.b64decode(_SAMPLE_EXCEL_B64)
    return send_file(
        io.BytesIO(data),
        as_attachment=True,
        download_name="main_qexcel.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@uploads_bp.route("/sample_excel_item_import")
def sample_excel2():
    import io
    from backend.sample_excel_b64_item_import import _SAMPLE_EXCEL_B642
    data = base64.b64decode(_SAMPLE_EXCEL_B642)
    return send_file(
        io.BytesIO(data),
        as_attachment=True,
        download_name="item_import.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
