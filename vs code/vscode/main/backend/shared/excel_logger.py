from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

EXCEL_FILE = None

def set_excel_path(path):
    global EXCEL_FILE
    EXCEL_FILE = path

def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Logs"
        ws.append(["Timestamp", "Level", "Message"])
        wb.save(EXCEL_FILE)

def write_log(level, message):
    if not EXCEL_FILE:
        return  # Reporting disabled

    init_excel()

    wb = load_workbook(EXCEL_FILE)

    if "Logs" not in wb.sheetnames:
        ws = wb.create_sheet("Logs")
        ws.append(["Timestamp", "Level", "Message"])
    else:
        ws = wb["Logs"]

    ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        str(level),
        str(message) if message else ""
    ])

    wb.save(EXCEL_FILE)
