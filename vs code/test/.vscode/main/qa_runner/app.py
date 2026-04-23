"""
app.py  -  Q-Audit Local Runner  (v9 — queue-order-first)

FIX: The terminal init log now prints the ACTUAL resolved active_modules
     list (in drag priority order) instead of cfg["modules"] which could
     be stale or differently ordered.
     execute() logs the order AFTER active_modules is fully built.
"""

import os
import sys
import json
import time
import base64
import threading
import subprocess
import sqlite3
import re
import hashlib
import secrets
import uuid

from functools import wraps
from io import BytesIO

from openpyxl import load_workbook

from flask import (
    Flask,
    render_template,
    request,
    jsonify,
    Response,
    stream_with_context,
    send_file,
    redirect,
    url_for,
    session,
    make_response,
)

import logger as _logger
from logger import log_buffer, log_lock
from config_builder import build_config_py
from sample_excel_b64 import _SAMPLE_EXCEL_B64
from sample_excel_b64_item_import import _SAMPLE_EXCEL_B642

# ---------------------------------------------------------------------------
# App
# ---------------------------------------------------------------------------
app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY") or secrets.token_hex(32)

_BASE = os.path.dirname(os.path.abspath(__file__))

UPLOAD_DIR = os.path.join(_BASE, "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

_ALLOWED_UPLOAD_EXTS = {".xlsx", ".xls"}

# ---------------------------------------------------------------------------
# Auth decorator
# ---------------------------------------------------------------------------
def _require_login(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("auth_login_page"))
        return f(*args, **kwargs)
    return wrapper

# ---------------------------------------------------------------------------
# Runtime state  (thread-safe)
# ---------------------------------------------------------------------------
running:        bool     = False
current_run_id: str|None = None
_state_lock               = threading.Lock()
_active_proc              = None

# ---------------------------------------------------------------------------
# Main DB
# ---------------------------------------------------------------------------
DB_PATH = os.path.join(_BASE, "qaudit.db")

def _init_db():
    with sqlite3.connect(DB_PATH) as con:
        con.execute("""
            CREATE TABLE IF NOT EXISTS runs (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                run_id     TEXT,
                step       TEXT,
                status     TEXT,
                progress   INTEGER DEFAULT 0,
                started_at REAL,
                ended_at   REAL
            )
        """)
        con.commit()

def _db_insert_step(run_id, step, status, progress):
    with sqlite3.connect(DB_PATH) as con:
        con.execute(
            "INSERT INTO runs(run_id,step,status,progress,started_at) VALUES(?,?,?,?,?)",
            (run_id, step, status, progress, time.time()),
        )
        con.commit()

def _db_finish_run(run_id, status):
    with sqlite3.connect(DB_PATH) as con:
        con.execute(
            "UPDATE runs SET ended_at=?, status=? WHERE run_id=? AND ended_at IS NULL",
            (time.time(), status, run_id),
        )
        con.commit()

def _get_history():
    with sqlite3.connect(DB_PATH) as con:
        rows = con.execute(
            "SELECT run_id, step, status, progress, started_at, ended_at "
            "FROM runs ORDER BY id DESC LIMIT 100"
        ).fetchall()
    return [
        {"run_id": r[0], "step": r[1], "status": r[2],
         "progress": r[3], "started_at": r[4], "ended_at": r[5]}
        for r in rows
    ]

_init_db()

# ---------------------------------------------------------------------------
# Auth DB
# ---------------------------------------------------------------------------
AUTH_DB_PATH = os.path.join(_BASE, "auth.db")

def _init_auth_db():
    with sqlite3.connect(AUTH_DB_PATH) as con:
        con.executescript("""
            CREATE TABLE IF NOT EXISTS users (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                email         TEXT UNIQUE NOT NULL,
                first_name    TEXT,
                last_name     TEXT,
                password_hash TEXT,
                is_verified   INTEGER DEFAULT 0,
                created_at    REAL,
                last_login    REAL
            );
            CREATE TABLE IF NOT EXISTS otp_tokens (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                email      TEXT NOT NULL,
                purpose    TEXT NOT NULL,
                code_hash  TEXT NOT NULL,
                expires_at REAL NOT NULL,
                used       INTEGER DEFAULT 0
            );
            CREATE INDEX IF NOT EXISTS idx_otp_email ON otp_tokens(email, purpose);
        """)
        con.commit()

_init_auth_db()

# ---------------------------------------------------------------------------
# Auth helpers
# ---------------------------------------------------------------------------
_EMAIL_RE       = re.compile(r"^[^\s@]+@[^\s@]+\.[^\s@]+$")
OTP_TTL_SECS    = 600
RESEND_COOLDOWN = 60
_last_otp: dict = {}


def _hash_password(pw: str) -> str:
    try:
        from argon2 import PasswordHasher
        return PasswordHasher().hash(pw)
    except ImportError:
        salt = secrets.token_hex(16)
        dk = hashlib.pbkdf2_hmac("sha256", pw.encode(), salt.encode(), 260_000)
        return f"pbkdf2${salt}${dk.hex()}"


def _verify_password(stored: str, pw: str) -> bool:
    if stored.startswith("$argon2"):
        from argon2 import PasswordHasher
        from argon2.exceptions import VerifyMismatchError
        try:
            PasswordHasher().verify(stored, pw)
            return True
        except VerifyMismatchError:
            return False
    if stored.startswith("pbkdf2$"):
        _, salt, dk_hex = stored.split("$")
        dk = hashlib.pbkdf2_hmac("sha256", pw.encode(), salt.encode(), 260_000)
        return secrets.compare_digest(dk.hex(), dk_hex)
    return False


def _issue_otp(email: str, purpose: str) -> str:
    code      = str(secrets.randbelow(900_000) + 100_000)
    code_hash = hashlib.sha256(code.encode()).hexdigest()
    with sqlite3.connect(AUTH_DB_PATH) as con:
        con.execute(
            "INSERT INTO otp_tokens(email,purpose,code_hash,expires_at) VALUES(?,?,?,?)",
            (email, purpose, code_hash, time.time() + OTP_TTL_SECS),
        )
        con.commit()
    return code


def _verify_otp(email: str, purpose: str, code: str) -> bool:
    code_hash = hashlib.sha256(code.encode()).hexdigest()
    with sqlite3.connect(AUTH_DB_PATH) as con:
        row = con.execute(
            "SELECT id FROM otp_tokens "
            "WHERE email=? AND purpose=? AND code_hash=? AND used=0 AND expires_at>?",
            (email, purpose, code_hash, time.time()),
        ).fetchone()
        if not row:
            return False
        con.execute("UPDATE otp_tokens SET used=1 WHERE id=?", (row[0],))
        con.commit()
    return True


def _get_resend_wait(email: str, purpose: str) -> float:
    with sqlite3.connect(AUTH_DB_PATH) as con:
        row = con.execute(
            "SELECT MAX(expires_at) FROM otp_tokens WHERE email=? AND purpose=?",
            (email, purpose),
        ).fetchone()
    if not row or not row[0]:
        return 0
    issued_at = row[0] - OTP_TTL_SECS
    return max(RESEND_COOLDOWN - (time.time() - issued_at), 0)


def _send_otp_email(to_email: str, code: str, purpose: str):
    label = "Verification" if purpose == "signup" else "Password Reset"
    print(f"\n  📧  {label} code for {to_email}: {code}\n", flush=True)
    _last_otp["code"]    = code
    _last_otp["email"]   = to_email
    _last_otp["purpose"] = purpose
    _last_otp["ts"]      = time.time()


def _json_ok(**kwargs):
    return jsonify({"ok": True, **kwargs})


def _json_err(msg: str, status: int = 400):
    return jsonify({"ok": False, "error": msg}), status


# ──────────────────────────────────────────────────────────────────────────
# DEV ROUTE — DELETE before deploying to production
# ──────────────────────────────────────────────────────────────────────────
@app.route("/dev/otp")
def dev_otp():
    if not _last_otp:
        return jsonify({"ok": False})
    return jsonify({"ok": True, **_last_otp})


# ──────────────────────────────────────────────────────────────────────────
# AUTH PAGE ROUTES
# ──────────────────────────────────────────────────────────────────────────
@app.route("/login")
def auth_login_page():
    return render_template("auth.html")

@app.route("/signup")
def auth_signup_page():
    return render_template("auth.html")

@app.route("/verify")
def auth_verify_page():
    return render_template("auth.html")

@app.route("/set-password")
def auth_setpw_page():
    return render_template("auth.html")

@app.route("/forgot-password")
def auth_forgot_page():
    return render_template("auth.html")

@app.route("/reset-password")
def auth_reset_page():
    return render_template("auth.html")


# ──────────────────────────────────────────────────────────────────────────
# AUTH API ROUTES
# ──────────────────────────────────────────────────────────────────────────

@app.route("/auth/signup", methods=["POST"])
def api_signup():
    data       = request.get_json() or {}
    email      = (data.get("email")      or "").strip().lower()
    first_name = (data.get("first_name") or "").strip()
    last_name  = (data.get("last_name")  or "").strip()

    if not email or not _EMAIL_RE.match(email):
        return _json_err("Invalid email address.")
    if not first_name or not last_name:
        return _json_err("First and last name are required.")

    with sqlite3.connect(AUTH_DB_PATH) as con:
        existing = con.execute(
            "SELECT id, is_verified FROM users WHERE email=?", (email,)
        ).fetchone()

    if existing:
        _, verified = existing
        if verified:
            return _json_err("An account with this email already exists.")

    with sqlite3.connect(AUTH_DB_PATH) as con:
        if not existing:
            con.execute(
                "INSERT INTO users(email,first_name,last_name,created_at) VALUES(?,?,?,?)",
                (email, first_name, last_name, time.time()),
            )
            con.commit()

    wait = _get_resend_wait(email, "signup")
    if wait > 0:
        return _json_err(f"Please wait {int(wait)+1}s before requesting another code.")

    code = _issue_otp(email, "signup")
    _send_otp_email(email, code, "signup")
    return _json_ok(message="Verification code sent.")


@app.route("/auth/verify", methods=["POST"])
def api_verify():
    data  = request.get_json() or {}
    email = (data.get("email") or "").strip().lower()
    code  = (data.get("code")  or "").strip()

    if not email or not code:
        return _json_err("Email and code are required.")
    if not _verify_otp(email, "signup", code):
        return _json_err("Invalid or expired verification code.")

    with sqlite3.connect(AUTH_DB_PATH) as con:
        con.execute("UPDATE users SET is_verified=1 WHERE email=?", (email,))
        con.commit()
    return _json_ok()


@app.route("/auth/resend", methods=["POST"])
def api_resend():
    data  = request.get_json() or {}
    email = (data.get("email") or "").strip().lower()

    if not email:
        return _json_err("Email is required.")

    wait = _get_resend_wait(email, "signup")
    if wait > 0:
        return _json_err(f"Please wait {int(wait)+1}s before resending.")

    code = _issue_otp(email, "signup")
    _send_otp_email(email, code, "signup")
    return _json_ok(message="Code resent.")


@app.route("/auth/set-password", methods=["POST"])
def api_set_password():
    data     = request.get_json() or {}
    email    = (data.get("email")    or "").strip().lower()
    password =  data.get("password") or ""

    if not email:
        return _json_err("Email is required.")
    if len(password) < 8:
        return _json_err("Password must be at least 8 characters.")

    with sqlite3.connect(AUTH_DB_PATH) as con:
        row = con.execute(
            "SELECT id, is_verified FROM users WHERE email=?", (email,)
        ).fetchone()
    if not row:
        return _json_err("Account not found.")
    if not row[1]:
        return _json_err("Email not verified. Please complete verification first.")

    pw_hash = _hash_password(password)
    with sqlite3.connect(AUTH_DB_PATH) as con:
        con.execute(
            "UPDATE users SET password_hash=? WHERE email=?", (pw_hash, email)
        )
        con.commit()

    session["user_id"] = row[0]
    session["email"]   = email
    return _json_ok(message="Account created successfully.")


@app.route("/auth/login", methods=["POST"])
def api_login():
    data     = request.get_json() or {}
    email    = (data.get("email")    or "").strip().lower()
    password =  data.get("password") or ""

    if not email or not password:
        return _json_err("Email and password are required.")

    with sqlite3.connect(AUTH_DB_PATH) as con:
        row = con.execute(
            "SELECT id, password_hash, is_verified FROM users WHERE email=?", (email,)
        ).fetchone()

    if not row:
        return _json_err("Invalid email or password.")

    uid, pw_hash, verified = row

    if not verified:
        return _json_err("Please verify your email before signing in.")
    if not pw_hash:
        return _json_err("Account setup incomplete. Please set a password.")
    if not _verify_password(pw_hash, password):
        return _json_err("Invalid email or password.")

    with sqlite3.connect(AUTH_DB_PATH) as con:
        con.execute("UPDATE users SET last_login=? WHERE id=?", (time.time(), uid))
        con.commit()

    session["user_id"] = uid
    session["email"]   = email
    return _json_ok()


@app.route("/auth/logout", methods=["POST"])
def api_logout():
    session.clear()
    return _json_ok(redirect="/login")


@app.route("/auth/forgot-password", methods=["POST"])
def api_forgot_password():
    data  = request.get_json() or {}
    email = (data.get("email") or "").strip().lower()

    if not email or not _EMAIL_RE.match(email):
        return _json_err("Invalid email address.")

    with sqlite3.connect(AUTH_DB_PATH) as con:
        row = con.execute(
            "SELECT id FROM users WHERE email=? AND is_verified=1", (email,)
        ).fetchone()

    if row:
        wait = _get_resend_wait(email, "reset")
        if wait <= 0:
            code = _issue_otp(email, "reset")
            _send_otp_email(email, code, "reset")

    return _json_ok(message="If that address is registered you will receive a reset code.")


@app.route("/auth/reset-password", methods=["POST"])
def api_reset_password():
    data     = request.get_json() or {}
    email    = (data.get("email")    or "").strip().lower()
    code     = (data.get("code")     or "").strip()
    password =  data.get("password") or ""

    if not email or not code:
        return _json_err("Email and code are required.")
    if len(password) < 8:
        return _json_err("Password must be at least 8 characters.")
    if not _verify_otp(email, "reset", code):
        return _json_err("Invalid or expired reset code.")

    pw_hash = _hash_password(password)
    with sqlite3.connect(AUTH_DB_PATH) as con:
        con.execute(
            "UPDATE users SET password_hash=? WHERE email=?", (pw_hash, email)
        )
        con.commit()
    return _json_ok(message="Password updated successfully.")


# ──────────────────────────────────────────────────────────────────────────
# MAIN APP ROUTES
# ──────────────────────────────────────────────────────────────────────────

@app.route("/")
@_require_login
def index():
    return render_template("index.html")


@app.route("/upload", methods=["POST"])
def upload():
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify({"error": "No file provided", "path": ""}), 400

    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in _ALLOWED_UPLOAD_EXTS:
        return jsonify({"error": f"File type '{ext}' not allowed. Use .xlsx or .xls", "path": ""}), 400

    safe_name = os.path.basename(f.filename)
    dest = os.path.join(UPLOAD_DIR, safe_name)
    f.save(dest)
    return jsonify({"path": dest})


@app.route("/inspect_audit_excel", methods=["POST"])
def inspect_audit_excel():
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify({"ok": False, "errors": ["No file provided."], "sheet_names": [], "headers_by_sheet": {}}), 400

    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in _ALLOWED_UPLOAD_EXTS:
        return jsonify({"ok": False, "errors": [f"File type '{ext}' not allowed. Use .xlsx or .xls"], "sheet_names": [], "headers_by_sheet": {}}), 400

    if ext == ".xls":
        return jsonify({
            "ok": False,
            "errors": [".xls inspection is not available on this machine. Please use .xlsx or enter the mapping manually."],
            "sheet_names": [],
            "headers_by_sheet": {},
        }), 400

    try:
        file_bytes = f.read()
        wb = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
        headers_by_sheet = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if row:
                headers_by_sheet[sheet_name] = [str(val).strip() for val in row if val is not None and str(val).strip()]
            else:
                headers_by_sheet[sheet_name] = []
        wb.close()
        default_sheet = "auditor_1" if "auditor_1" in headers_by_sheet else (wb.sheetnames[0] if wb.sheetnames else "")
        warnings = []
        if default_sheet and not headers_by_sheet.get(default_sheet):
            warnings.append(f"Sheet '{default_sheet}' has no detected header row.")
        return jsonify({
            "ok": True,
            "sheet_names": wb.sheetnames,
            "default_sheet": default_sheet,
            "headers_by_sheet": headers_by_sheet,
            "warnings": warnings,
            "errors": [],
        })
    except Exception as exc:
        return jsonify({
            "ok": False,
            "errors": [f"Unable to inspect the uploaded Excel: {exc}"],
            "sheet_names": [],
            "headers_by_sheet": {},
        }), 400


@app.route("/sample_excel")
def sample_excel():
    import io
    data = base64.b64decode(_SAMPLE_EXCEL_B64)
    return send_file(
        io.BytesIO(data),
        as_attachment=True,
        download_name="main_qexcel.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@app.route("/sample_excel_item_import")
def sample_excel2():
    import io
    data = base64.b64decode(_SAMPLE_EXCEL_B642)
    return send_file(
        io.BytesIO(data),
        as_attachment=True,
        download_name="item_import.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/run", methods=["POST"])
def run_route():
    global running, current_run_id, _active_proc

    with _state_lock:
        if running:
            return jsonify({"error": "A run is already in progress"}), 409
        running = True

    cfg = request.get_json()
    _logger.reset()

    rid = str(uuid.uuid4())[:8]
    current_run_id = rid

    # Normalize common boolean-ish strings coming from the frontend so that
    # module selection and Config.py generation stay consistent.
    def _is_true(v):
        if isinstance(v, bool):
            return v
        if isinstance(v, str):
            return v.strip().lower() in ("true", "1", "yes", "on")
        return bool(v)

    # ── Module catalogue: flag key → display name ──────────────────────────
    MODULE_KEYS = [
        ("run_q_setting",           "Q_Setting"),
        ("run_Q_SA",                "Q_SA"),
        ("run_Q_SA1",               "Q_SA1"),
        ("run_Q_as_table",          "Q_as_table"),
        ("run_Q_Recently_Audit",    "Q_Recently_Audit"),
        ("run_Q_audit_Summary",     "Q_audit_Summary"),   # idx 5
        ("run_create_audit",        "Audit_Plan"),         # idx 6
        ("run_create_Ad_hoc_Audit", "Ad_hoc_Audit"),      # idx 6 (alt)
        ("run_setup_Audit",         "Ongoing_Audit"),     # idx 7
        ("run_A_SA",                "A_SA"),              # idx 8
        ("run_A_SA2",               "A_SA2"),             # idx 9
        ("run_A_as_table",          "A_as_table"),        # idx 10
        ("run_A_audit_Summary",     "A_audit_Summary"),   # idx 11
        ("run_A_Recently_Audit",    "A_Recently_Audit"),  # idx 12
        ("run_create_group",        "Create_Group"),      # idx 13
        ("run_locatio_setup",       "Location_Setup"),    # idx 14
        ("run_import_app",          "Import_App"),         # idx 15
    ]

    name_to_flag = {name: key  for key, name in MODULE_KEYS}
    flag_to_name = {key:  name for key, name in MODULE_KEYS}

    # Coerce all module flags to real booleans early.
    for key, _ in MODULE_KEYS:
        if key in cfg:
            cfg[key] = _is_true(cfg.get(key))

    # ── Step A: Normalize Config ───────────────────────────────────────────
    # Normalise cfg.modules → boolean flags (backward compat) ───────────
    modules_list = cfg.get("modules")
    if modules_list and isinstance(modules_list, list):
        for item in modules_list:
            if isinstance(item, str) and item in name_to_flag:
                cfg[name_to_flag[item]] = True

    # ── UI module index → flag key ─────────────────────────────────────────
    idx_to_flag = {
        0:  "run_q_setting",
        1:  "run_Q_SA",
        2:  "run_Q_SA1",
        3:  "run_Q_as_table",
        4:  "run_Q_Recently_Audit",
        5:  "run_Q_audit_Summary",
        6:  "run_create_audit",      # resolved to Ad_hoc variant by _resolve_idx6_flag
        7:  "run_setup_Audit",
        8:  "run_A_SA",
        9:  "run_A_SA2",
        10: "run_A_as_table",
        11: "run_A_audit_Summary",
        12: "run_A_Recently_Audit",
        13: "run_create_group",
        14: "run_locatio_setup",
        15: "run_import_app",
    }
    flag_to_idx = {v: k for k, v in idx_to_flag.items()}

    # ── Special case: idx 6 can be either Audit_Plan OR Ad_hoc ────────────
    # If the frontend sent MODULE_RUN_ORDER containing 6, resolve which
    # sub-flag is actually on.
    def _resolve_idx6_flag(cfg: dict) -> str:
        """Return the correct flag for UI index 6 based on cfg sub-flags."""
        if cfg.get("run_create_Ad_hoc_Audit"):
            return "run_create_Ad_hoc_Audit"
        return "run_create_audit"

    # ── Step B: Resolve Active Modules ─────────────────────────────────────
    # Build active_modules in drag priority order ────────────────────────
    order = cfg.get("MODULE_RUN_ORDER")  # list of ints from queue.js

    active_modules: list[tuple[str, str]] = []

    if order and isinstance(order, list):
        seen: set[str] = set()
        for item in order:
            flag = None

            if isinstance(item, int):
                if item == 6:
                    flag = _resolve_idx6_flag(cfg)
                else:
                    flag = idx_to_flag.get(item)
            elif isinstance(item, str):
                flag = item if item.startswith("run_") else name_to_flag.get(item)

            if not flag or flag in seen:
                continue
            seen.add(flag)

            if cfg.get(flag):
                active_modules.append((flag, flag_to_name.get(flag, flag)))

        # Fallback: if drag order produced nothing, use static order
        if not active_modules:
            active_modules = [(k, n) for k, n in MODULE_KEYS if cfg.get(k)]
    else:
        # No drag order sent — static MODULE_KEYS order
        active_modules = [(k, n) for k, n in MODULE_KEYS if cfg.get(k)]

    total = max(len(active_modules), 1)

    # ── Execute ────────────────────────────────────────────────────────────
    def execute():
        global running, _active_proc
        try:
            # ┌─────────────────────────────────────────────────────────────┐
            # │  FIX: log the ACTUAL resolved order, not cfg["modules"]     │
            # │  This is what the terminal "Run #... started" line shows.   │
            # └─────────────────────────────────────────────────────────────┘
            _logger.log(json.dumps({
                "type":    "init",
                "run_id":  rid,
                "total":   total,
                "modules": [n for _, n in active_modules],   # ← resolved order
            }))

            # Option: run the full queue in a single subprocess so the same
            # browser session flows across modules (e.g., Q_setting â†’ Q_SA).
            # Default remains isolated per-module execution.
            # Default to single-process flow so selected modules run in one browser/session.
            # Set QA_RUNNER_SINGLE_PROCESS=0 to restore legacy per-module isolation.
            single_process = os.getenv("QA_RUNNER_SINGLE_PROCESS", "1").strip().lower() in ("1", "true", "yes", "on")

            if single_process:
                # ── Step C: Generate Config.py ────────────────────────────────────────
                cfg["MODULE_RUN_ORDER"] = [flag for flag, n in active_modules]
                config_path = os.path.join(_BASE, "Config.py")
                with open(config_path, "w", encoding="utf-8") as fp:
                    fp.write(build_config_py(cfg))

                _logger.log("\n[INFO] ▶ Launching: Full flow (single process)")
                _logger.log("─" * 52)

                # ── Step D: Execute Flow (Single Process Mode) ────────────────────────
                main_path  = os.path.join(_BASE, "main.py")
                launch_cmd = [sys.executable, "-u", main_path]

                env = dict(os.environ)
                env["PYTHONUNBUFFERED"] = "1"

                proc = subprocess.Popen(
                    launch_cmd,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    text=True,
                    encoding="utf-8",
                    errors="replace",
                    cwd=_BASE,
                    env=env,
                    bufsize=1,
                )

                with _state_lock:
                    _active_proc = proc

                for line in iter(proc.stdout.readline, ""):
                    stripped = line.rstrip()
                    _logger.log(stripped)

                proc.wait()

                if proc.returncode != 0:
                    _logger.log(f"❌ Aborted: flow failed (exit {proc.returncode})")
                    _db_finish_run(rid, "error")
                    return

                _db_insert_step(rid, "Full flow", "done", 100)
            else:
                done_count = 0

                for flag, mname in active_modules:

                    # Isolate just this module for Config.py (optimized)
                    module_cfg = dict(cfg)

                    # Turn off every other module flag
                    for k, _ in MODULE_KEYS:
                        if module_cfg.get(k):
                            module_cfg[k] = False

                    # Enable strictly the current module
                    module_cfg[flag] = True

                    # Tell main.py to run only this index
                    target_idx = flag_to_idx.get(flag)
                    module_cfg["MODULE_RUN_ORDER"] = [target_idx] if target_idx is not None else []

                    # ── Step C: Generate Config.py ────────────────────────────────────────
                    # Write Config.py
                    config_path = os.path.join(_BASE, "Config.py")
                    with open(config_path, "w", encoding="utf-8") as fp:
                        fp.write(build_config_py(module_cfg))

                    _logger.log(f"\n[INFO] ▶ Launching: {mname}")
                    _logger.log("─" * 52)

                    # ── Step D: Execute Flow (sub process) ────────────────────────────────
                    main_path  = os.path.join(_BASE, "main.py")
                    # Use unbuffered mode so subprocess output streams line-by-line
                    # into the web UI and terminal in real time.
                    launch_cmd = [sys.executable, "-u", main_path]

                    env = dict(os.environ)
                    env["PYTHONUNBUFFERED"] = "1"

                    proc = subprocess.Popen(
                        launch_cmd,
                        stdout=subprocess.PIPE,
                        stderr=subprocess.STDOUT,
                        text=True,
                        encoding="utf-8",
                        errors="replace",
                        cwd=_BASE,
                        env=env,
                        bufsize=1,
                    )

                    with _state_lock:
                        _active_proc = proc

                    module_failed = False

                    for line in iter(proc.stdout.readline, ""):
                        stripped = line.rstrip()
                        _logger.log(stripped)
                        sl = stripped.lower()
                        if ("playwright error" in sl
                                or "unexpected error" in sl
                                or ("error" in sl and "audit name already exists" in sl)):
                            module_failed = True

                proc.wait()

                if proc.returncode != 0 or module_failed:
                    _logger.log(f"❌ Aborted: {mname} failed (exit {proc.returncode})")
                    _db_finish_run(rid, "error")
                    return

                done_count = min(done_count + 1, total)
                pct = int(done_count / total * 100)

                _logger.log(json.dumps({
                    "type":     "progress",
                    "step":     mname,
                    "progress": pct,
                    "done":     done_count,
                    "total":    total,
                }))

                _db_insert_step(rid, mname, "done", pct)

            _logger.log("─" * 52)
            _logger.log("[OK] Run completed successfully.")
            _logger.log(json.dumps({
                "type":     "progress",
                "progress": 100,
                "done":     total,
                "total":    total,
            }))
            _db_finish_run(rid, "done")

        except Exception as e:
            _logger.log(f"[ERROR] {e}")
            _db_finish_run(rid, "error")
        finally:
            with _state_lock:
                running      = False
                _active_proc = None
            _logger.log("__DONE__")

    threading.Thread(target=execute, daemon=True).start()
    return jsonify({"status": "started", "run_id": rid, "total": total})


@app.route("/stop", methods=["POST"])
def stop():
    global running, _active_proc
    with _state_lock:
        proc = _active_proc

    if proc:
        try:
            proc.terminate()
            proc.wait(timeout=5)
        except subprocess.TimeoutExpired:
            proc.kill()
        except Exception:
            pass

    with _state_lock:
        running = False

    _logger.log("[WARN] Run stopped by user.")
    _logger.log("__ERROR__")
    return jsonify({"status": "stopped"})


@app.route("/stream")
def stream():
    def generate():
        sent = 0
        while True:
            with log_lock:
                chunk = log_buffer[sent:]
            for line in chunk:
                sent += 1
                yield f"data: {line}\n\n"
                if line in ("__DONE__", "__ERROR__"):
                    return
            time.sleep(0.01)

    return Response(stream_with_context(generate()), mimetype="text/event-stream")


@app.route("/history")
def history():
    return jsonify(_get_history())


# ──────────────────────────────────────────────────────────────────────────
# Entry point
# ──────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("\n  Q-Audit Runner  (v9)")
    print("  " + "─" * 57)
    print("  Open in browser →  http://localhost:5000")
    print("  Press Ctrl+C to stop\n")
    app.run(debug=False, port=5000, threaded=True)
